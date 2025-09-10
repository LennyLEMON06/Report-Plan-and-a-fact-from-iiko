import requests
import hashlib
from datetime import datetime
import json
import urllib3
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import timedelta
import calendar
import time

# Отключение предупреждений о сертификатах
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class IikoDataCollector:
    def __init__(self, base_url, login="angelinalina", password="092002"):
        self.base_url = base_url.strip()
        self.login = login
        self.password = password
        self.token = None
        self.session = requests.Session()
        self.session.verify = False

    def auth(self):
        """Аутентификация в системе iiko"""
        auth_url = f"{self.base_url}/auth"
        try:
            password_hash = hashlib.sha1(self.password.encode()).hexdigest()
            response = self.session.post(
                auth_url,
                data={'login': self.login, 'pass': password_hash},
                headers={'Content-Type': 'application/x-www-form-urlencoded'}
            )
            if response.status_code == 200:
                self.token = response.text.strip()
                print("✅ Авторизация успешна")
                return True
            print(f"❌ Ошибка авторизации: {response.status_code} - {response.text}")
            return False
        except Exception as e:
            print(f"❌ Ошибка подключения: {str(e)}")
            return False

    def get_report_data(self, preset_id, date_from, date_to, save_raw=False):
        """Получение данных отчета"""
        if not self.token and not self.auth():
            return None

        url = f"{self.base_url}/v2/reports/olap/byPresetId/{preset_id}"
        
        params = {
            'key': self.token,
            'dateFrom': date_from.strftime('%Y-%m-%d'),
            'dateTo': date_to.strftime('%Y-%m-%d')
        }

        try:
            print(f"Отправляем запрос в {self.base_url}...")
            response = self.session.get(url, params=params)
            
            if response.status_code == 200:
                print("✅ Данные успешно получены")
                json_data = response.json()
                
                if save_raw:
                    Path("raw_data").mkdir(exist_ok=True)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"raw_data/report_{timestamp}.json"
                    with open(filename, 'w', encoding='utf-8') as f:
                        json.dump(json_data, f, ensure_ascii=False, indent=2)
                    print(f"✅ Сырые данные сохранены в файл: {filename}")
                
                return json_data
            
            print(f"❌ Ошибка получения отчета: {response.status_code}")
            print(f"Ответ сервера: {response.text}")
            return None
            
        except Exception as e:
            print(f"❌ Ошибка при получении отчета: {str(e)}")
            return None

    def get_average_report_data(self, average_id, date_from, date_to):
        """Получение данных отчета по average_id"""
        if not self.token and not self.auth():
            return None

        # Используем average_id как preset_id для получения отчета
        url = f"{self.base_url}/v2/reports/olap/byPresetId/{average_id}"
        
        params = {
            'key': self.token,
            'dateFrom': date_from.strftime('%Y-%m-%d'),
            'dateTo': date_to.strftime('%Y-%m-%d')
        }

        try:
            response = self.session.get(url, params=params)
            
            if response.status_code == 200:
                print("✅ Данные average отчета успешно получены")
                return response.json()
            else:
                print(f"❌ Ошибка получения average отчета: {response.status_code}")
                print(f"Ответ сервера: {response.text}")
                return None
                
        except Exception as e:
            print(f"❌ Ошибка при получении average отчета: {str(e)}")
            return None

def load_bases_config():
    """Загрузка конфигурации баз из файла"""
    try:
        with open("bases_config.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"❌ Ошибка загрузки bases_config.json: {e}")
        return {}

def select_bases(bases):
    """Выбор баз для выгрузки"""
    print("\nДоступные базы:")
    keys = list(bases.keys())
    for i, name in enumerate(keys, 1):
        print(f"{i}. {name}")
    
    selected = input("\nВведите номера баз через запятую (например: 1,3,5) или 'all' для всех: ").strip()
    
    if selected.lower() == "all":
        return keys
    
    try:
        # Разбиваем ввод на числа и преобразуем в целые
        indices = [int(x.strip()) for x in selected.split(',') if x.strip()]
        
        # Проверяем корректность номеров
        invalid_indices = [i for i in indices if i < 1 or i > len(keys)]
        if invalid_indices:
            print(f"❌ Некорректные номера: {invalid_indices}")
            return []
        
        # Возвращаем выбранные базы
        return [keys[i-1] for i in indices]
    except ValueError:
        print("❌ Неверный ввод")
        return []

def select_sheets():
    """Выбор листов для выгрузки"""
    print("\nДоступные листы:")
    print("1. Свод План-Факт")
    print("2. План_факт наполняемость")
    print("3. План_факт стоимость блюд")
    print("4. Все листы")
    
    selected = input("\nВведите номера листов через запятую (например: 1,3) или выберите вариант (4 или 5): ").strip()
    
    if selected == "4":
        return [1, 2, 3]  # Все листы
    
    try:
        # Разбиваем ввод на числа и преобразуем в целые
        indices = [int(x.strip()) for x in selected.split(',') if x.strip()]
        
        # Проверяем корректность номеров
        invalid_indices = [i for i in indices if i < 1 or i > 3]
        if invalid_indices:
            print(f"❌ Некорректные номера листов: {invalid_indices}")
            return []
        
        return indices
    except ValueError:
        print("❌ Неверный ввод")
        return []

def create_main_report_sheet(wb, df, cash_registers):
    """Создание основного листа Свод План-Факт"""
    ws = wb.create_sheet(title="Свод План-Факт")

    # Заголовок
    ws['A1'] = "СВОДНЫЙ ОТЧЕТ ПО ВЫРУЧКЕ"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    # Цвета для разных баз
    base_colors = {
        0: "E2EFDA",  # Зеленый
        1: "FFEB9C",  # Желтый  
        2: "C6E0B4",  # Светло-зеленый
        3: "FFE699",  # Светло-желтый
        4: "A9D08E",  # Зеленый
        5: "F8CBAD",  # Оранжевый
        6: "9CC2E5",  # Голубой
        7: "D9E1F2",  # Светло-синий
    }

    # Заголовки для каждой кассы
    headers = ["День недели", "Дата", "План", "Факт", "% выполнения"]
    
    col_offset = 0
    base_index = 0
    for cash_register in cash_registers:
        start_col = col_offset + 1
        end_col = col_offset + len(headers)
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        ws.cell(row=2, column=start_col, value=cash_register)
        ws.cell(row=2, column=start_col).font = Font(bold=True)
        ws.cell(row=2, column=start_col).alignment = Alignment(horizontal='center')
        
        # Уникальный цвет для каждой базы
        color_index = base_index % len(base_colors)
        color = base_colors[color_index]
        ws.cell(row=2, column=start_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Заголовки столбцов
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=col_offset + i + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        col_offset += len(headers)
        base_index += 1

    # Заполняем данные
    dates = sorted(df['OpenDate.Typed'].unique())
    
    current_row = 4
    for date in dates:
        day_of_week = df[df['OpenDate.Typed'] == date]['DayOfWeekOpen'].iloc[0]
        formatted_date = df[df['OpenDate.Typed'] == date]['Date'].iloc[0]
        
        col_offset = 0
        for cash_register in cash_registers:
            ws.cell(row=current_row, column=col_offset + 1, value=day_of_week)
            ws.cell(row=current_row, column=col_offset + 2, value=formatted_date)
            
            # План (берем из данных или 0 если нет)
            cash_data = df[(df['OpenDate.Typed'] == date) & (df['Store.Name'] == cash_register)]
            plan_value = cash_data.iloc[0]['PlanValue'] if not cash_data.empty and 'PlanValue' in cash_data.iloc[0] else 0
            ws.cell(row=current_row, column=col_offset + 3, value=plan_value)
            
            # Факт
            if not cash_data.empty:
                fact_value = cash_data.iloc[0]['DishDiscountSumInt']
                ws.cell(row=current_row, column=col_offset + 4, value=fact_value)
                
                # Процент выполнения — ВСЕГДА формула
                plan_col = get_column_letter(col_offset + 3)
                fact_col = get_column_letter(col_offset + 4)
                formula = f"=IF({plan_col}{current_row}=0, 0, {fact_col}{current_row}/{plan_col}{current_row})"
                ws.cell(row=current_row, column=col_offset + 5, value=formula)
                ws.cell(row=current_row, column=col_offset + 5).number_format = '0.00%'
            
            col_offset += len(headers)
        
        current_row += 1

    # Итоговая строка
    total_row = current_row
    ws.cell(row=total_row, column=1, value="ИТОГО")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Дата выгрузки в столбце "Дата" для строки ИТОГО
    current_date = datetime.now().strftime('%d.%m.%Y')
    ws.cell(row=total_row, column=2, value=current_date)
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    col_offset = 0
    for cash_register in cash_registers:
        # Сумма плана
        plan_col = get_column_letter(col_offset + 3)
        ws.cell(row=total_row, column=col_offset + 3, value=f"=SUM({plan_col}4:{plan_col}{current_row-1})")
        ws.cell(row=total_row, column=col_offset + 3).font = Font(bold=True)
        
        # Сумма факта
        fact_col = get_column_letter(col_offset + 4)
        ws.cell(row=total_row, column=col_offset + 4, value=f"=SUM({fact_col}4:{fact_col}{current_row-1})")
        ws.cell(row=total_row, column=col_offset + 4).font = Font(bold=True)
        
        # Процент выполнения (факт/план для итоговой строки)
        plan_total_col = get_column_letter(col_offset + 3)
        fact_total_col = get_column_letter(col_offset + 4)
        formula = f"=IF({plan_total_col}{current_row}=0, 0, {fact_total_col}{current_row}/{plan_total_col}{current_row})"
        ws.cell(row=total_row, column=col_offset + 5, value=formula)
        ws.cell(row=total_row, column=col_offset + 5).number_format = '0.00%'
        ws.cell(row=total_row, column=col_offset + 5).font = Font(bold=True)
        
        col_offset += len(headers)

    # Форматирование
    for col in range(1, col_offset + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in range(2, total_row + 1):
        for col in range(1, col_offset + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Условное форматирование для процентов выполнения
    col_offset = 0
    for cash_register in cash_registers:
        percent_col = col_offset + 5
        percent_letter = get_column_letter(percent_col)
        percent_range = f"{percent_letter}4:{percent_letter}{current_row-1}"

        # Зеленый для >= 100%
        rule_green = CellIsRule(
            operator='greaterThanOrEqual',
            formula=['1'],
            stopIfTrue=False,
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        )
        ws.conditional_formatting.add(percent_range, rule_green)
        
        # Красный для < 100%
        rule_red = CellIsRule(
            operator='lessThan',
            formula=['1'],
            stopIfTrue=False,
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        )
        ws.conditional_formatting.add(percent_range, rule_red)
        
        col_offset += len(headers)

def create_occupancy_report_sheet(wb, df, cash_registers, bases_config, average_data_dict):
    """Создание листа План_факт наполняемость"""
    ws = wb.create_sheet(title="План_факт наполняемость")

    # Заголовок
    ws['A1'] = "ПЛАН-ФАКТ НАПОЛНЯЕМОСТЬ"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    # Цвета для разных баз
    base_colors = {
        0: "E2EFDA",  # Зеленый
        1: "FFEB9C",  # Желтый  
        2: "C6E0B4",  # Светло-зеленый
        3: "FFE699",  # Светло-желтый
        4: "A9D08E",  # Зеленый
        5: "F8CBAD",  # Оранжевый
        6: "9CC2E5",  # Голубой
        7: "D9E1F2",  # Светло-синий
    }

    # Заголовки для каждой кассы
    headers = ["День недели", "Дата", "План", "Факт", "% выполнения"]
    
    col_offset = 0
    base_index = 0
    for cash_register in cash_registers:
        start_col = col_offset + 1
        end_col = col_offset + len(headers)
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        ws.cell(row=2, column=start_col, value=cash_register)
        ws.cell(row=2, column=start_col).font = Font(bold=True)
        ws.cell(row=2, column=start_col).alignment = Alignment(horizontal='center')
        
        # Уникальный цвет для каждой базы
        color_index = base_index % len(base_colors)
        color = base_colors[color_index]
        ws.cell(row=2, column=start_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Заголовки столбцов
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=col_offset + i + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        col_offset += len(headers)
        base_index += 1

    # Заполняем данные (наполняемость)
    dates = sorted(df['OpenDate.Typed'].unique())
    
    current_row = 4
    for date in dates:
        day_of_week = df[df['OpenDate.Typed'] == date]['DayOfWeekOpen'].iloc[0]
        formatted_date = df[df['OpenDate.Typed'] == date]['Date'].iloc[0]
        
        col_offset = 0
        for cash_register in cash_registers:
            ws.cell(row=current_row, column=col_offset + 1, value=day_of_week)
            ws.cell(row=current_row, column=col_offset + 2, value=formatted_date)
            
            # Находим соответствующую базу для получения average_id
            cash_data = df[(df['OpenDate.Typed'] == date) & (df['Store.Name'] == cash_register)]
            base_name = cash_data.iloc[0]['BaseName'] if not cash_data.empty else None
            
            # План наполняемости (берем из данных или 0 если нет)
            plan_occupancy = cash_data.iloc[0].get('PlanOccupancy', 0) if not cash_data.empty else 0
            ws.cell(row=current_row, column=col_offset + 3, value=plan_occupancy)
            
            # Факт наполняемости (берем из average отчета)
            fact_occupancy = 0
            if base_name and base_name in average_data_dict:
                avg_data = average_data_dict[base_name]
                # Ищем данные за нужную дату
                date_str = date.strftime('%Y-%m-%d')
                for item in avg_data:
                    if item.get('OpenDate.Typed', '').startswith(date_str):
                        dish_amount = item.get('DishAmountInt', 0)
                        guest_num = item.get('GuestNum', 0)
                        if guest_num > 0:
                            fact_occupancy = round(dish_amount / guest_num, 2)
                        break
            
            ws.cell(row=current_row, column=col_offset + 4, value=fact_occupancy)
            
            # Процент выполнения — ВСЕГДА формула (даже если план = 0)
            plan_col_letter = get_column_letter(col_offset + 3)
            fact_col_letter = get_column_letter(col_offset + 4)
            formula = f"=IF({plan_col_letter}{current_row}=0, 0, {fact_col_letter}{current_row}/{plan_col_letter}{current_row})"
            ws.cell(row=current_row, column=col_offset + 5, value=formula)
            ws.cell(row=current_row, column=col_offset + 5).number_format = '0.00%'
            
            col_offset += len(headers)
        
        current_row += 1

    # Итоговая строка
    total_row = current_row
    ws.cell(row=total_row, column=1, value="ИТОГО")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Дата выгрузки в столбце "Дата" для строки ИТОГО
    current_date = datetime.now().strftime('%d.%m.%Y')
    ws.cell(row=total_row, column=2, value=current_date)
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    col_offset = 0
    for cash_register in cash_registers:
        # Среднее плана (округленное до 2 знаков)
        plan_col = get_column_letter(col_offset + 3)
        ws.cell(row=total_row, column=col_offset + 3, value=f"=ROUND(AVERAGE({plan_col}4:{plan_col}{current_row-1}),2)")
        ws.cell(row=total_row, column=col_offset + 3).font = Font(bold=True)
        
        # Среднее факта (округленное до 2 знаков)
        fact_col = get_column_letter(col_offset + 4)
        ws.cell(row=total_row, column=col_offset + 4, value=f"=ROUND(AVERAGE({fact_col}4:{fact_col}{current_row-1}),2)")
        ws.cell(row=total_row, column=col_offset + 4).font = Font(bold=True)
        
        # Процент выполнения (факт/план для итоговой строки)
        plan_total_col = get_column_letter(col_offset + 3)
        fact_total_col = get_column_letter(col_offset + 4)
        formula = f"=IF({plan_total_col}{current_row}=0, 0, {fact_total_col}{current_row}/{plan_total_col}{current_row})"
        ws.cell(row=total_row, column=col_offset + 5, value=formula)
        ws.cell(row=total_row, column=col_offset + 5).number_format = '0.00%'
        ws.cell(row=total_row, column=col_offset + 5).font = Font(bold=True)
        
        col_offset += len(headers)

    # Форматирование
    for col in range(1, col_offset + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in range(2, total_row + 1):
        for col in range(1, col_offset + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Условное форматирование для процентов выполнения
    col_offset = 0
    for cash_register in cash_registers:
        percent_col = col_offset + 5
        percent_letter = get_column_letter(percent_col)
        percent_range = f"{percent_letter}4:{percent_letter}{current_row-1}"

        # Зеленый для >= 100%
        rule_green = CellIsRule(
            operator='greaterThanOrEqual',
            formula=['1'],
            stopIfTrue=False,
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        )
        ws.conditional_formatting.add(percent_range, rule_green)
        
        # Красный для < 100%
        rule_red = CellIsRule(
            operator='lessThan',
            formula=['1'],
            stopIfTrue=False,
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        )
        ws.conditional_formatting.add(percent_range, rule_red)
        
        col_offset += len(headers)

def create_dish_price_report_sheet(wb, df, cash_registers, bases_config, average_data_dict):
    """Создание листа План_факт стоимость блюд"""
    ws = wb.create_sheet(title="План_факт стоимость блюд")

    # Заголовок
    ws['A1'] = "ПЛАН-ФАКТ СТОИМОСТЬ БЛЮД"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    # Цвета для разных баз
    base_colors = {
        0: "E2EFDA",  # Зеленый
        1: "FFEB9C",  # Желтый  
        2: "C6E0B4",  # Светло-зеленый
        3: "FFE699",  # Светло-желтый
        4: "A9D08E",  # Зеленый
        5: "F8CBAD",  # Оранжевый
        6: "9CC2E5",  # Голубой
        7: "D9E1F2",  # Светло-синий
    }

    # Заголовки для каждой кассы
    headers = ["День недели", "Дата", "План", "Факт", "% выполнения"]
    
    col_offset = 0
    base_index = 0
    for cash_register in cash_registers:
        start_col = col_offset + 1
        end_col = col_offset + len(headers)
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        ws.cell(row=2, column=start_col, value=cash_register)
        ws.cell(row=2, column=start_col).font = Font(bold=True)
        ws.cell(row=2, column=start_col).alignment = Alignment(horizontal='center')
        
        # Уникальный цвет для каждой базы
        color_index = base_index % len(base_colors)
        color = base_colors[color_index]
        ws.cell(row=2, column=start_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Заголовки столбцов
        for i, header in enumerate(headers):
            cell = ws.cell(row=3, column=col_offset + i + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        col_offset += len(headers)
        base_index += 1

    # Заполняем данные (стоимость блюд)
    dates = sorted(df['OpenDate.Typed'].unique())
    
    current_row = 4
    for date in dates:
        day_of_week = df[df['OpenDate.Typed'] == date]['DayOfWeekOpen'].iloc[0]
        formatted_date = df[df['OpenDate.Typed'] == date]['Date'].iloc[0]
        
        col_offset = 0
        for cash_register in cash_registers:
            ws.cell(row=current_row, column=col_offset + 1, value=day_of_week)
            ws.cell(row=current_row, column=col_offset + 2, value=formatted_date)
            
            # Находим соответствующую базу для получения average_id
            cash_data = df[(df['OpenDate.Typed'] == date) & (df['Store.Name'] == cash_register)]
            base_name = cash_data.iloc[0]['BaseName'] if not cash_data.empty else None
            
            # План стоимости блюд (берем из данных или 0 если нет)
            plan_dish_price = cash_data.iloc[0].get('PlanDishPrice', 0) if not cash_data.empty else 0
            ws.cell(row=current_row, column=col_offset + 3, value=plan_dish_price)
            
            # Факт стоимости блюд (берем из average отчета)
            fact_dish_price = 0
            if base_name and base_name in average_data_dict:
                avg_data = average_data_dict[base_name]
                # Ищем данные за нужную дату
                date_str = date.strftime('%Y-%m-%d')
                for item in avg_data:
                    if item.get('OpenDate.Typed', '').startswith(date_str):
                        dish_discount_sum = item.get('DishDiscountSumInt', 0)
                        dish_amount = item.get('DishAmountInt', 0)
                        if dish_amount > 0:
                            fact_dish_price = round(dish_discount_sum / dish_amount, 2)
                        break
            
            ws.cell(row=current_row, column=col_offset + 4, value=fact_dish_price)
            
            # Процент выполнения — ВСЕГДА формула (даже если план = 0)
            plan_col_letter = get_column_letter(col_offset + 3)
            fact_col_letter = get_column_letter(col_offset + 4)
            formula = f"=IF({plan_col_letter}{current_row}=0, 0, {fact_col_letter}{current_row}/{plan_col_letter}{current_row})"
            ws.cell(row=current_row, column=col_offset + 5, value=formula)
            ws.cell(row=current_row, column=col_offset + 5).number_format = '0.00%'
            
            col_offset += len(headers)
        
        current_row += 1

    # Итоговая строка
    total_row = current_row
    ws.cell(row=total_row, column=1, value="ИТОГО")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Дата выгрузки в столбце "Дата" для строки ИТОГО
    current_date = datetime.now().strftime('%d.%m.%Y')
    ws.cell(row=total_row, column=2, value=current_date)
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    col_offset = 0
    for cash_register in cash_registers:
        # Среднее плана (округленное до 2 знаков)
        plan_col = get_column_letter(col_offset + 3)
        ws.cell(row=total_row, column=col_offset + 3, value=f"=ROUND(AVERAGE({plan_col}4:{plan_col}{current_row-1}),2)")
        ws.cell(row=total_row, column=col_offset + 3).font = Font(bold=True)
        
        # Среднее факта (округленное до 2 знаков)
        fact_col = get_column_letter(col_offset + 4)
        ws.cell(row=total_row, column=col_offset + 4, value=f"=ROUND(AVERAGE({fact_col}4:{fact_col}{current_row-1}),2)")
        ws.cell(row=total_row, column=col_offset + 4).font = Font(bold=True)
        
        # Процент выполнения (факт/план для итоговой строки)
        plan_total_col = get_column_letter(col_offset + 3)
        fact_total_col = get_column_letter(col_offset + 4)
        formula = f"=IF({plan_total_col}{current_row}=0, 0, {fact_total_col}{current_row}/{plan_total_col}{current_row})"
        ws.cell(row=total_row, column=col_offset + 5, value=formula)
        ws.cell(row=total_row, column=col_offset + 5).number_format = '0.00%'
        ws.cell(row=total_row, column=col_offset + 5).font = Font(bold=True)
        
        col_offset += len(headers)

    # Форматирование
    for col in range(1, col_offset + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    for row in range(2, total_row + 1):
        for col in range(1, col_offset + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Условное форматирование для процентов выполнения
    col_offset = 0
    for cash_register in cash_registers:
        percent_col = col_offset + 5
        percent_letter = get_column_letter(percent_col)
        percent_range = f"{percent_letter}4:{percent_letter}{current_row-1}"

        # Зеленый для >= 100%
        rule_green = CellIsRule(
            operator='greaterThanOrEqual',
            formula=['1'],
            stopIfTrue=False,
            fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        )
        ws.conditional_formatting.add(percent_range, rule_green)
        
        # Красный для < 100%
        rule_red = CellIsRule(
            operator='lessThan',
            formula=['1'],
            stopIfTrue=False,
            fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        )
        ws.conditional_formatting.add(percent_range, rule_red)
        
        col_offset += len(headers)

def create_excel_report(all_data, bases_config, output_filename, report_month, report_date, average_data_dict, selected_sheets):
    """Создание Excel отчета с выбранными листами"""
    if not all_data:
        print("❌ Нет данных для создания отчета")
        return False
    
    # Создаем DataFrame из всех данных
    df = pd.DataFrame(all_data)
    
    # Преобразуем дату в правильный формат
    df['OpenDate.Typed'] = pd.to_datetime(df['OpenDate.Typed'])
    df['Date'] = df['OpenDate.Typed'].dt.strftime('%d.%m.%Y')
    df['DayOfWeek'] = df['DayOfWeekOpen'].str.extract(r'(\d+)\.')[0].astype(int)
    
    # Сортируем по дате и дню недели
    df = df.sort_values(['OpenDate.Typed', 'DayOfWeek'])
    
    # Получаем уникальные кассы
    cash_registers = df['Store.Name'].unique()
    
    # Создаем Excel файл
    wb = Workbook()
    
    # Удаляем все существующие листы
    while len(wb.sheetnames) > 0:
        del wb[wb.sheetnames[0]]
    
    # Создаем выбранные листы
    if 1 in selected_sheets:
        create_main_report_sheet(wb, df, cash_registers)
    if 2 in selected_sheets:
        create_occupancy_report_sheet(wb, df, cash_registers, bases_config, average_data_dict)
    if 3 in selected_sheets:
        create_dish_price_report_sheet(wb, df, cash_registers, bases_config, average_data_dict)

    # Сохраняем файл
    wb.save(output_filename)
    print(f"✅ Отчет сохранен: {output_filename}")
    return True

def main():
    print("=== Получение и преобразование данных из iiko ===")
    
    # Загружаем конфиг
    bases = load_bases_config()
    if not bases:
        print("❌ Не удалось загрузить конфигурацию баз")
        return

    # Выбираем базы
    selected_bases = select_bases(bases)
    if not selected_bases:
        print("❌ Не выбрано ни одной базы")
        return

    # Выбираем листы
    selected_sheets = select_sheets()
    if not selected_sheets:
        print("❌ Не выбрано ни одного листа")
        return

    # Устанавливаем период
    year = 2025
    month = 9
    first_day = 1
    last_day = calendar.monthrange(year, month)[1]
    date_from = datetime(year, month, first_day)
    date_to = datetime(year, month, last_day) + timedelta(days=1)

    all_data = []
    average_data_dict = {}
    
    # Получаем данные из каждой базы
    for base_name in selected_bases:
        config = bases[base_name]
        print(f"\n=== Работаем с базой: {base_name} ===")
        
        collector = IikoDataCollector(config["url"])
        
        # Получаем основные данные
        data = collector.get_report_data(config["preset_id"], date_from, date_to)
        
        if data and 'data' in data:
            # Добавляем имя базы к каждому элементу
            for item in data['data']:
                item['BaseName'] = base_name
            all_data.extend(data['data'])
        
        # Получаем данные по average_id
        avg_data = collector.get_average_report_data(config["average_id"], date_from, date_to)
        if avg_data and 'data' in avg_data:
            average_data_dict[base_name] = avg_data['data']

    if all_data:
        # Создаем папку для отчетов
        Path("reports").mkdir(exist_ok=True)
        
        # Формируем имя файла в формате "Свод_План_Факт_09_01_09_2025_10_25"
        current_date = datetime.now()
        report_date_str = current_date.strftime("%d_%m_%Y")
        timestamp = int(time.time())
        excel_filename = f"reports/Свод_План_Факт_{month:02d}_{report_date_str}_{timestamp}.xlsx"
        
        # Создаем отчет
        create_excel_report(all_data, bases, excel_filename, month, current_date, average_data_dict, selected_sheets)
        
        print(f"\n✅ Отчет создан: {excel_filename}")
    else:
        print("❌ Нет данных для создания отчета")

    input("\nНажмите Enter для выхода...")

if __name__ == "__main__":
    main()