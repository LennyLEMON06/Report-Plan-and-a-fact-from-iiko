# report_generator.py — логика создания Excel-файлов

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime

class ReportGenerator:
    def __init__(self, log_callback=None):
        self.log = log_callback or (lambda msg: print(msg))

    def calculate_revenue_fact(self, cash_data, avg_data_dict, date_str, base_name):
        if not cash_data.empty:
            return cash_data.iloc[0].get('DishDiscountSumInt', 0)
        return 0

    def calculate_occupancy_fact(self, cash_data, avg_data_dict, date_str, base_name):
        fact_occupancy = 0
        if base_name and base_name in avg_data_dict:
            avg_data = avg_data_dict[base_name]
            for item in avg_data:
                if ('OpenDate.Typed' in item and 
                    item.get('OpenDate.Typed', '').startswith(date_str) and
                    item.get('Store.Name') == cash_data.iloc[0]['Store.Name'] if not cash_data.empty else None):
                    dish_amount = item.get('DishAmountInt', 0)
                    guest_num = item.get('GuestNum', 0)
                    if guest_num > 0:
                        fact_occupancy = round(dish_amount / guest_num, 2)
                    break
        return fact_occupancy

    def calculate_dish_price_fact(self, cash_data, avg_data_dict, date_str, base_name):
        fact_dish_price = 0
        if base_name and base_name in avg_data_dict:
            avg_data = avg_data_dict[base_name]
            for item in avg_data:
                if ('OpenDate.Typed' in item and 
                    item.get('OpenDate.Typed', '').startswith(date_str) and
                    item.get('Store.Name') == cash_data.iloc[0]['Store.Name'] if not cash_data.empty else None):
                    dish_discount_sum = item.get('DishDiscountSumInt', 0)
                    dish_amount = item.get('DishAmountInt', 0)
                    if dish_amount > 0:
                        fact_dish_price = round(dish_discount_sum / dish_amount, 2)
                    break
        return fact_dish_price

    def calculate_guest_flow_fact(self, cash_data, avg_data_dict, date_str, base_name):
        fact_guest_flow = 0
        if base_name and base_name in avg_data_dict:
            avg_data = avg_data_dict[base_name]
            for item in avg_data:
                if ('OpenDate.Typed' in item and 
                    item.get('OpenDate.Typed', '').startswith(date_str) and
                    item.get('Store.Name') == cash_data.iloc[0]['Store.Name'] if not cash_data.empty else None):
                    fact_guest_flow = item.get('GuestNum', 0)
                    break
        return fact_guest_flow

    def create_generic_report_sheet(self, wb, df, cash_registers, average_data_dict, sheet_title, fact_calculator, plan_field_name):
        ws = wb.create_sheet(title=sheet_title)
        ws['A1'] = sheet_title.upper()
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')

        base_colors = {
            0: "E2EFDA", 1: "FFEB9C", 2: "C6E0B4", 3: "FFE699",
            4: "A9D08E", 5: "F8CBAD", 6: "9CC2E5", 7: "D9E1F2",
        }

        headers = ["День недели", "Дата", "План", "Факт", "% выполнения"]
        col_offset = 0
        base_index = 0

        for cash_register in cash_registers:
            start_col = col_offset + 1
            end_col = col_offset + len(headers)
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            ws.cell(row=2, column=start_col, value=cash_register).font = Font(bold=True)
            ws.cell(row=2, column=start_col).alignment = Alignment(horizontal='center')
            color = base_colors[base_index % len(base_colors)]
            ws.cell(row=2, column=start_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            for i, header in enumerate(headers):
                cell = ws.cell(row=3, column=col_offset + i + 1, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            col_offset += len(headers)
            base_index += 1

        dates = sorted(df['OpenDate.Typed'].unique())
        current_row = 4

        for date in dates:
            day_of_week = df[df['OpenDate.Typed'] == date]['DayOfWeekOpen'].iloc[0] if not df[df['OpenDate.Typed'] == date].empty else ""
            formatted_date = df[df['OpenDate.Typed'] == date]['Date'].iloc[0] if not df[df['OpenDate.Typed'] == date].empty else ""
            col_offset = 0

            for cash_register in cash_registers:
                ws.cell(row=current_row, column=col_offset + 1, value=day_of_week)
                ws.cell(row=current_row, column=col_offset + 2, value=formatted_date)

                cash_data = df[(df['OpenDate.Typed'] == date) & (df['Store.Name'] == cash_register)]
                base_name = cash_data.iloc[0]['BaseName'] if not cash_data.empty and 'BaseName' in cash_data.iloc[0] else None

                plan_value = cash_data.iloc[0].get(plan_field_name, 0) if not cash_data.empty else 0
                ws.cell(row=current_row, column=col_offset + 3, value=plan_value)

                date_str = date.strftime('%Y-%m-%d')
                fact_value = fact_calculator(cash_data, average_data_dict, date_str, base_name)
                ws.cell(row=current_row, column=col_offset + 4, value=fact_value)

                plan_col_letter = get_column_letter(col_offset + 3)
                fact_col_letter = get_column_letter(col_offset + 4)
                formula = f"=IF({plan_col_letter}{current_row}=0, 0, {fact_col_letter}{current_row}/{plan_col_letter}{current_row})"
                ws.cell(row=current_row, column=col_offset + 5, value=formula).number_format = '0.00%'

                col_offset += len(headers)

            current_row += 1

        total_row = current_row
        ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=datetime.now().strftime('%d.%m.%Y')).font = Font(bold=True)

        col_offset = 0
        for cash_register in cash_registers:
            plan_col = get_column_letter(col_offset + 3)
            fact_col = get_column_letter(col_offset + 4)

            if "выручк" in sheet_title.lower() or "гостепоток" in sheet_title.lower():
                ws.cell(row=total_row, column=col_offset + 3, value=f"=SUM({plan_col}4:{plan_col}{current_row-1})")
            else:
                ws.cell(row=total_row, column=col_offset + 3, value=f"=ROUND(AVERAGE({plan_col}4:{plan_col}{current_row-1}),2)")

            if "выручк" in sheet_title.lower() or "гостепоток" in sheet_title.lower():
                ws.cell(row=total_row, column=col_offset + 4, value=f"=SUM({fact_col}4:{fact_col}{current_row-1})")
            else:
                ws.cell(row=total_row, column=col_offset + 4, value=f"=ROUND(AVERAGE({fact_col}4:{fact_col}{current_row-1}),2)")

            ws.cell(row=total_row, column=col_offset + 3).font = Font(bold=True)
            ws.cell(row=total_row, column=col_offset + 4).font = Font(bold=True)

            plan_total_col = get_column_letter(col_offset + 3)
            fact_total_col = get_column_letter(col_offset + 4)
            formula = f"=IF({plan_total_col}{current_row}=0, 0, {fact_total_col}{current_row}/{plan_total_col}{current_row})"
            ws.cell(row=total_row, column=col_offset + 5, value=formula).number_format = '0.00%'.font = Font(bold=True)

            col_offset += len(headers)

        for col in range(1, col_offset + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in range(2, total_row + 1):
            for col in range(1, col_offset + 1):
                ws.cell(row=row, column=col).border = thin_border

        col_offset = 0
        for cash_register in cash_registers:
            percent_col = col_offset + 5
            percent_letter = get_column_letter(percent_col)
            percent_range = f"{percent_letter}4:{percent_letter}{current_row-1}"

            rule_green = CellIsRule(operator='greaterThanOrEqual', formula=['1'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
            ws.conditional_formatting.add(percent_range, rule_green)

            rule_red = CellIsRule(operator='lessThan', formula=['1'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
            ws.conditional_formatting.add(percent_range, rule_red)

            col_offset += len(headers)

    def create_excel_report(self, all_data, bases_config, output_filename, start_date, end_date, average_data_dict, selected_sheets):
        if not all_data:
            self.log("❌ Нет данных для создания отчета")
            return False

        try:
            df = pd.DataFrame(all_data)
            df['OpenDate.Typed'] = pd.to_datetime(df['OpenDate.Typed'])
            df['Date'] = df['OpenDate.Typed'].dt.strftime('%d.%m.%Y')

            if 'DayOfWeekOpen' in df.columns:
                df['DayOfWeek'] = df['DayOfWeekOpen'].str.extract(r'(\d+)\.')[0].astype(int)
            else:
                df['DayOfWeek'] = df['OpenDate.Typed'].dt.dayofweek + 1

            df = df.sort_values(['OpenDate.Typed', 'DayOfWeek'])
            cash_registers = df['Store.Name'].unique() if 'Store.Name' in df.columns else []

            if len(cash_registers) == 0:
                self.log("❌ Нет данных о кассах")
                return False

            wb = Workbook()
            while len(wb.sheetnames) > 0:
                del wb[wb.sheetnames[0]]

            if 1 in selected_sheets:
                self.create_generic_report_sheet(
                    wb, df, cash_registers, average_data_dict,
                    "Свод План-Факт", self.calculate_revenue_fact, 'PlanValue'
                )
            if 2 in selected_sheets:
                self.create_generic_report_sheet(
                    wb, df, cash_registers, average_data_dict,
                    "План_факт наполняемость", self.calculate_occupancy_fact, 'PlanOccupancy'
                )
            if 3 in selected_sheets:
                self.create_generic_report_sheet(
                    wb, df, cash_registers, average_data_dict,
                    "План_факт стоимость блюд", self.calculate_dish_price_fact, 'PlanDishPrice'
                )
            if 4 in selected_sheets:
                self.create_generic_report_sheet(
                    wb, df, cash_registers, average_data_dict,
                    "План_факт гостепоток", self.calculate_guest_flow_fact, 'PlanGuestFlow'
                )

            wb.save(output_filename)
            self.log(f"✅ Отчет сохранен: {output_filename}")
            return True

        except Exception as e:
            self.log(f"❌ Ошибка при создании отчета: {str(e)}")
            return False